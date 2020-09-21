import openpyxl

try:
    fullinfo_file = openpyxl.load_workbook("Full_info.xlsx")
    sheet = fullinfo_file.active
    max_rows = sheet.max_row
    max_cols = sheet.max_column

    for row in range(1,max_rows):
        for col in range(1,max_cols+1):
            file_name = sheet.cell(row=1, column=col).value
            with open(file_name, "a", encoding="utf8") as txt_file:
                value = sheet.cell(row=row+1, column=col).value
                txt_file.write(f"{value}\n")

except FileNotFoundError as error:
    print(error)