import openpyxl
from openpyxl.styles import Font
import pyinputplus as pyip 

print("\n=========Start============\n")
num = pyip.inputInt("Enter Multiplication Value: ", greaterThan=0)

f = openpyxl.Workbook()
sheet = f.active
# every time this loop completes it add one row header and one column header
# and all columns multiplication for one row
for row in range(1,num+1):
    #set header values for rows
    sheet.cell(row=row+1, column=1, value=row).font = Font(bold=True)
    #set header values for columns
    sheet.cell(row=1, column=row+1, value=row).font = Font(bold=True)
    #every time this loop completes it set all the multiplication of row and
    #column for one row
    for col in range(1, num+1):
        #set multipliction value fro every cell
        sheet.cell(row=row+1, column=col+1, value=(row*col))

f.save("file.xlsx")
print("\n==========Success============\n")