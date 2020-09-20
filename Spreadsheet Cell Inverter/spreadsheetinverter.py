import openpyxl

f = openpyxl.load_workbook("file.xlsx")
main_sheet = f.active
main_sheet.title = "Main"
main_rows = main_sheet.max_row
main_cols = main_sheet.max_column
#new sheet
new_sheet = f.create_sheet("new_sheet")

#every time this loop completes it change on row data to columns in the 
#new sheet
for row in range(1, main_rows+1):
    for col in range(1,main_cols+1):
        value = main_sheet.cell(row=row, column=col).value
        new_sheet.cell(row=col, column=row, value=value)

f.save("file.xlsx")