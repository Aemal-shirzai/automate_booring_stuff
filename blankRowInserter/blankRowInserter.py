import openpyxl
import pyinputplus as pyip

f = openpyxl.load_workbook("file.xlsx")
main_sheet = f.active
main_sheet.title = "Main"
main_max_rows = main_sheet.max_row
main_max_cols = main_sheet.max_column

print("\n============== Start =================\n")
starting_empty_row = pyip.inputInt("Enter Starting Row: \
    (should not exceed main file rows): ", min=0, max=main_max_rows
    )
num_of_emtpy_row = pyip.inputInt("Enter # of empty rows: ",min=0)

if starting_empty_row <= main_max_rows:
    # 1- create new sheet to add the updated data
    new_sheet = f.create_sheet("new_sheet") 
    # 2- loop through every row 
    for row in range(1, main_max_rows+1):
        #3- loop through every column
        for col in range(1,main_max_cols+1):
            #4- read data from main sheet
            value = main_sheet.cell(row=row, column=col).value
            if row >= starting_empty_row:
                #6- add rows to new sheet which are commming after the spliting row
                # row + num_of_empty_row: add blank rows
                new_sheet.cell(row=row+num_of_emtpy_row, column=col, value=value)
            else:
                #5- add rows to new sheet which are comming before the spliting row
                new_sheet.cell(row=row, column=col, value=value)
    f.save("file.xlsx")
    print("\n======= Success: New sheet created in same file==========\n")
else:
    print("=========== Warning Can not add empty rows at the end ===========")
